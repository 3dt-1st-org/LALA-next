from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Sequence

DEFAULT_SOURCE_NAME = "data_portal"
DETAIL_DATASET_NAME = "경기도_카드 소비 데이터"
AGGREGATE_DATASET_NAME = "경기도_데이터분석 카드매출 시군구 성연령별 집계"
DEFAULT_VISITOR_TYPE = "domestic"
SUPPORTED_SUFFIXES = {".csv", ".xlsx"}

GYEONGGI_SIGUNGU_CODE_TO_NAME = {
    "41110": "수원시",
    "41111": "수원시",
    "41113": "수원시",
    "41115": "수원시",
    "41117": "수원시",
    "41130": "성남시",
    "41131": "성남시",
    "41133": "성남시",
    "41135": "성남시",
    "41150": "의정부시",
    "41170": "안양시",
    "41171": "안양시",
    "41173": "안양시",
    "41190": "부천시",
    "41192": "부천시",
    "41194": "부천시",
    "41196": "부천시",
    "41210": "광명시",
    "41220": "평택시",
    "41250": "동두천시",
    "41270": "안산시",
    "41271": "안산시",
    "41273": "안산시",
    "41280": "고양시",
    "41281": "고양시",
    "41285": "고양시",
    "41287": "고양시",
    "41290": "과천시",
    "41310": "구리시",
    "41360": "남양주시",
    "41370": "오산시",
    "41390": "시흥시",
    "41410": "군포시",
    "41430": "의왕시",
    "41450": "하남시",
    "41460": "용인시",
    "41461": "용인시",
    "41463": "용인시",
    "41465": "용인시",
    "41480": "파주시",
    "41500": "이천시",
    "41550": "안성시",
    "41570": "김포시",
    "41590": "화성시",
    "41610": "광주시",
    "41630": "양주시",
    "41650": "포천시",
    "41670": "여주시",
    "41800": "연천군",
    "41820": "가평군",
    "41830": "양평군",
}

MONTH_ALIASES = ("month", "base_month", "base_ym", "기준년월", "년월")
DATE_ALIASES = ("date", "base_date", "base_ymd", "기준년월일", "기준일자", "일자")
REGION_NAME_ALIASES = (
    "region_name_ko",
    "region_name",
    "sigungu_name",
    "sgg_nm",
    "시군구명",
    "시군명",
    "시군구",
    "시군",
    "행정동명",
)
REGION_CODE_ALIASES = (
    "region_code",
    "sigungu_code",
    "sgg_cd",
    "adm_cd",
    "시군구코드",
    "시군코드",
    "행정동코드",
)
INDUSTRY_CODE_ALIASES = (
    "industry_code",
    "업종코드",
    "카드사업종분류코드",
    "카드사 업종분류코드",
    "중분류업종코드",
)
INDUSTRY_NAME_ALIASES = (
    "industry_name_ko",
    "industry_name",
    "업종명",
    "카드사업종중분류명",
    "카드사 업종중분류명",
    "중분류업종명",
)
GENDER_ALIASES = ("gender", "sex", "성별")
AGE_GROUP_ALIASES = ("age_group", "age", "연령별", "연령대")
GENDER_AGE_CODE_ALIASES = ("성연령코드", "gender_age_code", "sex_age_code")
SPEND_AMOUNT_ALIASES = (
    "spend_amount",
    "sales_amount",
    "amount",
    "매출금액",
    "카드매출금액",
)
TRANSACTION_COUNT_ALIASES = (
    "transaction_count",
    "sales_count",
    "count",
    "매출건수",
    "결제건수",
)


@dataclass(frozen=True)
class CardSpendingAreaMonthly:
    month: date
    region_name_ko: str
    industry_code: str | None
    industry_name_ko: str | None
    spend_amount: Decimal | None
    transaction_count: int | None
    visitor_type: str
    primary_source: str

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "month": self.month.isoformat(),
            "region_name_ko": self.region_name_ko,
            "industry_code": self.industry_code,
            "industry_name_ko": self.industry_name_ko,
            "spend_amount": _decimal_text(self.spend_amount),
            "transaction_count": self.transaction_count,
            "visitor_type": self.visitor_type,
            "primary_source": self.primary_source,
        }


@dataclass(frozen=True)
class CardSpendingDemographic:
    month: date
    region_name_ko: str
    industry_code: str | None
    gender: str | None
    age_group: str | None
    spend_amount: Decimal | None
    transaction_count: int | None
    primary_source: str

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "month": self.month.isoformat(),
            "region_name_ko": self.region_name_ko,
            "industry_code": self.industry_code,
            "gender": self.gender,
            "age_group": self.age_group,
            "spend_amount": _decimal_text(self.spend_amount),
            "transaction_count": self.transaction_count,
            "primary_source": self.primary_source,
        }


@dataclass(frozen=True)
class CardSpendingParseResult:
    source_name: str
    dataset_name: str
    file_name: str
    file_sha256: str
    local_path: str
    input_row_count: int
    parsed_row_count: int
    skipped_row_count: int
    area_monthly_rows: tuple[CardSpendingAreaMonthly, ...]
    demographic_rows: tuple[CardSpendingDemographic, ...]
    warnings: tuple[str, ...]

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "source_name": self.source_name,
            "dataset_name": self.dataset_name,
            "file_name": self.file_name,
            "file_sha256": self.file_sha256,
            "input_row_count": self.input_row_count,
            "parsed_row_count": self.parsed_row_count,
            "skipped_row_count": self.skipped_row_count,
            "area_monthly_row_count": len(self.area_monthly_rows),
            "demographic_row_count": len(self.demographic_rows),
            "warnings": list(self.warnings),
            "preview": {
                "area_monthly": [
                    row.to_public_dict() for row in self.area_monthly_rows[:5]
                ],
                "demographics": [
                    row.to_public_dict() for row in self.demographic_rows[:5]
                ],
            },
        }


def parse_card_spending_file(
    *,
    path: Path | str,
    source_name: str = DEFAULT_SOURCE_NAME,
    dataset_name: str = DETAIL_DATASET_NAME,
    visitor_type: str = DEFAULT_VISITOR_TYPE,
    region_code_map: dict[str, str] | None = None,
    row_limit: int | None = None,
) -> CardSpendingParseResult:
    source_path = Path(path)
    if not source_path.exists():
        raise FileNotFoundError(f"Card spending file does not exist: {source_path}")
    if source_path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError("Card spending file must be CSV or XLSX.")
    if row_limit is not None and row_limit <= 0:
        raise ValueError("row_limit must be positive when provided.")

    mapping = dict(GYEONGGI_SIGUNGU_CODE_TO_NAME)
    if region_code_map:
        mapping.update({_normalize_region_code(k): v for k, v in region_code_map.items()})

    rows = _read_rows(source_path)
    if row_limit is not None:
        rows = rows[:row_limit]

    area_groups: dict[tuple[Any, ...], dict[str, Any]] = defaultdict(
        lambda: {"spend_amount": None, "transaction_count": None}
    )
    demo_groups: dict[tuple[Any, ...], dict[str, Any]] = defaultdict(
        lambda: {"spend_amount": None, "transaction_count": None}
    )
    warning_messages: list[str] = []
    skipped = 0
    parsed = 0

    for row in rows:
        normalized = _normalize_row(row)
        parsed_row = _parse_row(
            normalized=normalized,
            source_name=source_name,
            visitor_type=visitor_type,
            region_code_map=mapping,
        )
        if parsed_row is None:
            skipped += 1
            _append_warning_once(warning_messages, "Some rows were skipped because month, region, or sales metrics were missing.")
            continue

        parsed += 1
        _merge_group(
            area_groups[
                (
                    parsed_row["month"],
                    parsed_row["region_name_ko"],
                    parsed_row["industry_code"],
                    parsed_row["industry_name_ko"],
                    parsed_row["visitor_type"],
                    parsed_row["primary_source"],
                )
            ],
            parsed_row["spend_amount"],
            parsed_row["transaction_count"],
        )

        if parsed_row["gender"] or parsed_row["age_group"]:
            _merge_group(
                demo_groups[
                    (
                        parsed_row["month"],
                        parsed_row["region_name_ko"],
                        parsed_row["industry_code"],
                        parsed_row["gender"],
                        parsed_row["age_group"],
                        parsed_row["primary_source"],
                    )
                ],
                parsed_row["spend_amount"],
                parsed_row["transaction_count"],
            )

    area_rows = tuple(
        CardSpendingAreaMonthly(
            month=key[0],
            region_name_ko=key[1],
            industry_code=key[2],
            industry_name_ko=key[3],
            visitor_type=key[4],
            primary_source=key[5],
            spend_amount=value["spend_amount"],
            transaction_count=value["transaction_count"],
        )
        for key, value in sorted(area_groups.items(), key=_sortable_group_key)
    )
    demo_rows = tuple(
        CardSpendingDemographic(
            month=key[0],
            region_name_ko=key[1],
            industry_code=key[2],
            gender=key[3],
            age_group=key[4],
            primary_source=key[5],
            spend_amount=value["spend_amount"],
            transaction_count=value["transaction_count"],
        )
        for key, value in sorted(demo_groups.items(), key=_sortable_group_key)
    )

    return CardSpendingParseResult(
        source_name=source_name,
        dataset_name=dataset_name,
        file_name=source_path.name,
        file_sha256=_sha256_file(source_path),
        local_path=str(source_path),
        input_row_count=len(rows),
        parsed_row_count=parsed,
        skipped_row_count=skipped,
        area_monthly_rows=area_rows,
        demographic_rows=demo_rows,
        warnings=tuple(warning_messages),
    )


def load_region_code_map(path: Path | str) -> dict[str, str]:
    mapping_path = Path(path)
    if not mapping_path.exists():
        raise FileNotFoundError(f"Region map file does not exist: {mapping_path}")

    result: dict[str, str] = {}
    for row in _read_rows(mapping_path):
        normalized = _normalize_row(row)
        code = _pick(normalized, REGION_CODE_ALIASES + ("code",))
        name = _pick(normalized, REGION_NAME_ALIASES + ("name",))
        if code and name:
            result[_normalize_region_code(code)] = name.strip()
    return result


def insert_card_spending_result(
    *,
    dsn: str,
    result: CardSpendingParseResult,
    connect_timeout: int,
) -> dict[str, Any]:
    import psycopg2

    source_sql = """
        INSERT INTO ingest.source_files (
            source_name,
            dataset_name,
            file_name,
            file_sha256,
            local_path
        )
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id
    """
    duplicate_sql = """
        SELECT id
        FROM ingest.source_files
        WHERE source_name = %s
          AND dataset_name = %s
          AND file_sha256 = %s
        ORDER BY downloaded_at DESC
        LIMIT 1
    """
    area_sql = """
        INSERT INTO economy.card_spending_area_monthly (
            month,
            region_name_ko,
            industry_code,
            industry_name_ko,
            spend_amount,
            transaction_count,
            visitor_type,
            primary_source,
            source_file_id
        )
        VALUES (
            %(month)s,
            %(region_name_ko)s,
            %(industry_code)s,
            %(industry_name_ko)s,
            %(spend_amount)s,
            %(transaction_count)s,
            %(visitor_type)s,
            %(primary_source)s,
            %(source_file_id)s
        )
    """
    demographic_sql = """
        INSERT INTO economy.card_spending_demographics (
            month,
            region_name_ko,
            industry_code,
            gender,
            age_group,
            spend_amount,
            transaction_count,
            primary_source,
            source_file_id
        )
        VALUES (
            %(month)s,
            %(region_name_ko)s,
            %(industry_code)s,
            %(gender)s,
            %(age_group)s,
            %(spend_amount)s,
            %(transaction_count)s,
            %(primary_source)s,
            %(source_file_id)s
        )
    """

    with psycopg2.connect(dsn, connect_timeout=connect_timeout) as conn:
        with conn.cursor() as cur:
            cur.execute(
                duplicate_sql,
                (result.source_name, result.dataset_name, result.file_sha256),
            )
            duplicate = cur.fetchone()
            if duplicate:
                return {
                    "ok": True,
                    "skipped_duplicate": True,
                    "source_file_id": str(duplicate[0]),
                    "inserted_area_monthly_rows": 0,
                    "inserted_demographic_rows": 0,
                }

            cur.execute(
                source_sql,
                (
                    result.source_name,
                    result.dataset_name,
                    result.file_name,
                    result.file_sha256,
                    result.local_path,
                ),
            )
            source_file_id = str(cur.fetchone()[0])

            inserted_area = 0
            for row in result.area_monthly_rows:
                params = row.to_public_dict()
                params["month"] = row.month
                params["spend_amount"] = row.spend_amount
                params["source_file_id"] = source_file_id
                cur.execute(area_sql, params)
                inserted_area += cur.rowcount

            inserted_demographics = 0
            for row in result.demographic_rows:
                params = row.to_public_dict()
                params["month"] = row.month
                params["spend_amount"] = row.spend_amount
                params["source_file_id"] = source_file_id
                cur.execute(demographic_sql, params)
                inserted_demographics += cur.rowcount
        conn.commit()

    return {
        "ok": True,
        "skipped_duplicate": False,
        "source_file_id": source_file_id,
        "inserted_area_monthly_rows": inserted_area,
        "inserted_demographic_rows": inserted_demographics,
    }


def _parse_row(
    *,
    normalized: dict[str, str],
    source_name: str,
    visitor_type: str,
    region_code_map: dict[str, str],
) -> dict[str, Any] | None:
    month = _parse_month(_pick(normalized, MONTH_ALIASES))
    if month is None:
        month = _parse_month(_pick(normalized, DATE_ALIASES))

    region_name = _pick(normalized, REGION_NAME_ALIASES)
    if not region_name:
        region_code = _pick(normalized, REGION_CODE_ALIASES)
        region_name = _region_name_from_code(region_code, region_code_map)

    spend_amount = _parse_decimal(_pick(normalized, SPEND_AMOUNT_ALIASES))
    transaction_count = _parse_int(_pick(normalized, TRANSACTION_COUNT_ALIASES))

    if month is None or not region_name or (spend_amount is None and transaction_count is None):
        return None

    gender = _normalize_gender(_pick(normalized, GENDER_ALIASES))
    age_group = _normalize_age_group(_pick(normalized, AGE_GROUP_ALIASES))
    combined_gender, combined_age = _parse_gender_age_code(
        _pick(normalized, GENDER_AGE_CODE_ALIASES)
    )
    gender = gender or combined_gender
    age_group = age_group or combined_age

    return {
        "month": month,
        "region_name_ko": region_name.strip(),
        "industry_code": _optional_text(_pick(normalized, INDUSTRY_CODE_ALIASES)),
        "industry_name_ko": _optional_text(_pick(normalized, INDUSTRY_NAME_ALIASES)),
        "gender": gender,
        "age_group": age_group,
        "spend_amount": spend_amount,
        "transaction_count": transaction_count,
        "visitor_type": visitor_type,
        "primary_source": source_name,
    }


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix == ".xlsx":
        return _read_xlsx_rows(path)
    raise ValueError("Unsupported file suffix.")


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [dict(row) for row in csv.DictReader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header: tuple[Any, ...] | None = None
    for raw in rows:
        if raw and any(cell is not None and str(cell).strip() for cell in raw):
            header = raw
            break
    if header is None:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in header]
    result: list[dict[str, Any]] = []
    for raw in rows:
        if not raw or not any(cell is not None and str(cell).strip() for cell in raw):
            continue
        result.append(
            {
                headers[index]: value
                for index, value in enumerate(raw[: len(headers)])
                if headers[index]
            }
        )
    return result


def _normalize_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        _normalize_key(key): "" if value is None else str(value).strip()
        for key, value in row.items()
        if key is not None and str(key).strip()
    }


def _normalize_key(value: Any) -> str:
    return re.sub(r"[\s_\-./()·:]+", "", str(value).strip()).lower()


def _pick(row: dict[str, str], aliases: Sequence[str]) -> str | None:
    for alias in aliases:
        value = row.get(_normalize_key(alias))
        if value:
            return value
    return None


def _parse_month(value: Any) -> date | None:
    text = _optional_text(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8:
        digits = digits[:6]
    if len(digits) != 6:
        return None
    try:
        return date(int(digits[:4]), int(digits[4:6]), 1)
    except ValueError:
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    text = _optional_text(value)
    if not text:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _parse_int(value: Any) -> int | None:
    decimal = _parse_decimal(value)
    if decimal is None:
        return None
    return int(decimal)


def _normalize_gender(value: Any) -> str | None:
    text = _optional_text(value)
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"f", "female", "woman", "women", "여", "여성"}:
        return "female"
    if lowered in {"m", "male", "man", "men", "남", "남성"}:
        return "male"
    return text


def _normalize_age_group(value: Any) -> str | None:
    text = _optional_text(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if digits:
        return f"{int(digits)}s"
    return text


def _parse_gender_age_code(value: Any) -> tuple[str | None, str | None]:
    text = _optional_text(value)
    if not text:
        return None, None
    match = re.match(r"^\s*([FMfm])\D*(\d{1,3})", text)
    if not match:
        return None, None
    return _normalize_gender(match.group(1)), _normalize_age_group(match.group(2))


def _region_name_from_code(
    value: Any,
    region_code_map: dict[str, str],
) -> str | None:
    code = _normalize_region_code(value)
    if not code:
        return None
    for candidate in (code, code[:5], code[:4], code[:2]):
        if candidate in region_code_map:
            return region_code_map[candidate]
    return None


def _normalize_region_code(value: Any) -> str:
    return re.sub(r"\D", "", str(value or "").strip())


def _merge_group(
    group: dict[str, Any],
    spend_amount: Decimal | None,
    transaction_count: int | None,
) -> None:
    if spend_amount is not None:
        group["spend_amount"] = (group["spend_amount"] or Decimal("0")) + spend_amount
    if transaction_count is not None:
        group["transaction_count"] = (group["transaction_count"] or 0) + transaction_count


def _sortable_group_key(item: tuple[tuple[Any, ...], dict[str, Any]]) -> tuple[Any, ...]:
    return tuple("" if value is None else value for value in item[0])


def _sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _append_warning_once(messages: list[str], message: str) -> None:
    if message not in messages:
        messages.append(message)


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value, "f")
